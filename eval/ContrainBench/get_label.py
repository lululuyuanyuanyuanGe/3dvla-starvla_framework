import random
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 定义物品和容器（按要求的顺序）
containers = [
    "棕色台子", "白色果篮", "棕色方形果篮", "紫色果盘", "咖啡色箱子"
]
items = [
    "奶牛玩具", "河马玩具", "大象玩具", "奥利奥", "茄子", "葡萄",
    "白色运动饮料瓶子", "橙色瓶子", "橙子", "小绿色薯片", "桃子"
]

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "任务数据"

# 写入表头
ws.append(["任务序号", "需要存在的物体", "任务"])

# 生成300组任务数据
tasks = []
prev_toy = None
random.seed(42)  # 确保结果可复现

def is_valid_selection(selected_items, prev_toy):
    """检查选择是否有效：不重复物品、最多一个玩具、不重复prev_toy"""
    # 检查是否有重复物品
    if len(set(selected_items)) < 3:
        return False
    
    # 检查是否同时出现多个玩具
    toy_count = sum(1 for item in selected_items if "玩具" in item)
    if toy_count > 1:
        return False
    
    # 检查是否重复prev_toy
    if prev_toy and prev_toy in selected_items:
        return False
    
    return True

for task_id in range(1, 301):
    valid_selection = False
    selected_items = []
    
    # 尝试找到有效的物品组合
    for _ in range(100):  # 最多尝试100次
        # 决定是否包含玩具（50%概率）
        include_toy = random.random() > 0.8
        
        if include_toy and prev_toy is None:
            # 可以包含任意玩具
            candidates = items.copy()
        elif include_toy:
            # 只能包含非prev_toy的玩具
            toy_options = [item for item in items 
                          if "玩具" in item and item != prev_toy]
            if not toy_options:
                include_toy = False
                candidates = [item for item in items if "玩具" not in item]
            else:
                # 随机选择一个非prev_toy的玩具
                selected_toy = random.choice(toy_options)
                other_items = [item for item in items 
                              if item != selected_toy and "玩具" not in item]
                # 选择2个非玩具物品
                selected_items = [selected_toy] + random.sample(other_items, 2)
                if is_valid_selection(selected_items, prev_toy):
                    valid_selection = True
                    break
                else:
                    continue
        else:
            # 不包含玩具
            candidates = [item for item in items if "玩具" not in item]
        
        # 随机选择3个不重复的物品
        selected_items = random.sample(candidates, 3)
        if is_valid_selection(selected_items, prev_toy):
            valid_selection = True
            break
    
    if not valid_selection:
        # 如果无法找到有效组合，强制使用非玩具物品
        non_toy_items = [item for item in items if "玩具" not in item]
        selected_items = random.sample(non_toy_items, 3)
    
    # 更新玩具记录
    toy_in_group = next((item for item in selected_items if "玩具" in item), None)
    prev_toy = toy_in_group
    
    # 选择容器（1-3个）
    container_count = random.choice([1, 2, 3])
    
    if container_count == 1:
        container = random.choice(containers)
        selected_containers = [container] * 3
    elif container_count == 2:
        container_pair = random.sample(containers, 2)
        if random.random() > 0.5:
            selected_containers = [container_pair[0], container_pair[0], container_pair[1]]
        else:
            selected_containers = [container_pair[0], container_pair[1], container_pair[1]]
    else:
        selected_containers = random.sample(containers, 3)
    
    # 构建"需要存在的物体"字符串（容器在前，物品在后，都排序）
    used_containers = sorted(list(set(selected_containers)), key=lambda x: containers.index(x))
    used_items = sorted(list(set(selected_items)), key=lambda x: items.index(x))
    required_objects = ", ".join(used_containers + used_items)
    
    # 保存任务数据
    tasks.append({
        "task_id": task_id,
        "required_objects": required_objects,
        "actions": [
            f"{selected_items[0]}→{selected_containers[0]}",
            f"{selected_items[1]}→{selected_containers[1]}",
            f"{selected_items[2]}→{selected_containers[2]}"
        ]
    })

# 按"需要存在的物体"排序任务
tasks.sort(key=lambda x: x["required_objects"])

# 写入Excel并合并单元格
for task in tasks:
    start_row = ws.max_row + 1
    
    # 写入三行数据
    for i in range(3):
        ws.append([
            task["task_id"] if i == 0 else "",
            task["required_objects"] if i == 0 else "",
            task["actions"][i]
        ])
    
    # 合并单元格
    ws.merge_cells(f"A{start_row}:A{start_row+2}")
    ws.merge_cells(f"B{start_row}:B{start_row+2}")
    
    # 设置垂直居中
    for row in ws.iter_rows(min_row=start_row, max_row=start_row+2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

# 设置列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 25

# 保存Excel文件
wb.save("task_data_final_2.xlsx")
print("Excel文件已生成: task_data_final_2.xlsx")